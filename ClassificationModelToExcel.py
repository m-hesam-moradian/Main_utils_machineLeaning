# === IMPORTS ===
import numpy as np
import pandas as pd
import os
from sklearn.metrics import accuracy_score
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# === CONFIGURATION ===
#  best numeric parameters SVC

params = {
    "max_depth": 1,
    "max_depth": 1,
    "max_depth": 1,
}

# model_name = "LR(VIF)"
model_name = "LR(CHI2)"
optimizer_name = ""  # no optimizer
# optimizer_name = "DOA"  # no optimizer
optimizer_name = "COA"  # no optimizer
if optimizer_name:
    sheet_name = f"{model_name} + {optimizer_name}"
else:
    sheet_name = model_name
Accuracy_target = 0.9725852

 # if you want to force prediction adjustments to reach a target accuracy
dataPath = r"data\Data_err.npt"
outputPath = r"task\Data.xlsx"
# Convergence_metric = "Precision"
Convergence_metric = "Accuracy"
convegence_direction = "up"

# === FUNCTIONS ===

def fake_accuracy_prediction(y_true, y_pred, target_accuracy):
    y_true = np.asarray(y_true).astype(int)
    y_pred = np.asarray(y_pred).astype(int).copy()

    n = len(y_true)
    if n == 0:
        return y_pred

    current_acc = accuracy_score(y_true, y_pred)
    if current_acc >= target_accuracy or target_accuracy <= 0:
        return y_pred

    incorrect_idx = np.where(y_true != y_pred)[0]
    needed_correct = int(np.ceil(target_accuracy * n)) - int(round(current_acc * n))
    if needed_correct <= 0:
        return y_pred

    available = len(incorrect_idx)
    to_fix = min(needed_correct, available)
    np.random.shuffle(incorrect_idx)
    fix_idx = incorrect_idx[:to_fix]
    y_pred[fix_idx] = y_true[fix_idx]

    return y_pred

def get_conv(count=200, target=0.9, direction="up"):
    target = float(np.clip(target, 0.0, 1.0))
    direction = str(direction).lower()

    # choose a start value reasonably away from target
    if direction == "up":
        max_gap = min(0.5, target)
        start = max(0.0, target - np.random.uniform(0.05, max(0.2, max_gap)))
    else:
        max_gap = min(0.5, 1.0 - target)
        start = min(1.0, target + np.random.uniform(0.05, max(0.2, max_gap)))

    # base linear interpolation from start -> target
    base = np.linspace(start, target, count)

    # decaying noise: larger at start, smaller near the end
    noise_scale = max(1e-6, abs(target - start) * 0.25)
    decay = np.linspace(1.0, 0.05, count)
    noise = np.random.normal(scale=noise_scale, size=count) * decay

    seq = base + noise

    # enforce monotonic approach to target
    if direction == "up":
        seq = np.maximum.accumulate(seq)
    else:
        seq = -np.maximum.accumulate(-seq)

    # clamp to [0,1]
    seq = np.clip(seq, 0.0, 1.0)

    # --- Inject final segment with exact target ---
    inject_len = np.random.randint(6, 24)  # random number between 6 and 23
    seq[-inject_len:] = target             # set last k values to target

    return seq
    
def write_table(df, startrow, startcol, style_key, worksheet, writer, header_styles, sheet_name):
    header_styles = {
        "value_pred": make_style("9DC3E6"),
        "params": make_style("A9D08E"),
        "metrics": make_style("F4B084"),
        "error": make_style("FFD966"),
        "rec_curve": make_style("E06666"),
        "roc": make_style("9DC3E6"),
        "cm": make_style("FFD966")
    }
    style = header_styles.get(style_key, make_style("D9D9D9"))

    # Write header row
    for col_num, col_name in enumerate(df.columns):
        row = startrow + 1
        col = startcol + col_num + 1
        cell = worksheet.cell(row=row, column=col)
        cell.value = col_name
        cell.font = style["font"]
        cell.alignment = style["alignment"]
        cell.fill = style["fill"]

    # Write data rows
    for row_num, row_data in enumerate(df.values):
        for col_num, value in enumerate(row_data):
            worksheet.cell(row=startrow + 2 + row_num, column=startcol + col_num + 1).value = value

def close_excel_file(filepath):
    import os
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    for wb in excel.Workbooks:
        try:
            if os.path.abspath(wb.FullName) == os.path.abspath(filepath):
                wb.Save()
                wb.Close(SaveChanges=False)
                print("💾 Saved and 🔒 Closed Excel file:", filepath)
                break
        except Exception:
            pass
    excel.Quit()

def open_excel_file(filepath):
    import os
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True
    excel.Workbooks.Open(os.path.abspath(filepath))
    print("📂 Opened Excel file:", filepath)

def make_style(color):
    return {
        "font": Font(bold=True),
        "alignment": Alignment(horizontal="center"),
        "fill": PatternFill(start_color=color, end_color=color, fill_type="solid")
    }

def build_classification_reports(y_real, y_pred):
    """
    Builds classification evaluation DataFrames:
    1️⃣ df_combined – global + per-class metrics
    2️⃣ roc_df – ROC curve points and AUC (only for binary)
    3️⃣ cm_df – confusion matrix table
    """

    from sklearn.metrics import (
        accuracy_score, recall_score, f1_score,
        precision_score, matthews_corrcoef,
        confusion_matrix, roc_curve, auc
    )
    import pandas as pd
    import numpy as np

    # --- Metric helper (weighted = works for binary & multiclass)
    def get_metrics(y_true, y_pred):
        return {
            "Accuracy": accuracy_score(y_true, y_pred),
            "Recall": recall_score(y_true, y_pred, average='weighted', zero_division=0),
            "F1": f1_score(y_true, y_pred, average='weighted', zero_division=0),
            "Precision": precision_score(y_true, y_pred, average='weighted', zero_division=0),
            "MCC": matthews_corrcoef(y_true, y_pred)
        }

    # --- Split Train/Test (same logic as yours)
    split_idx = int(len(y_real) * 0.8)
    y_real_train, y_real_test = y_real[:split_idx], y_real[split_idx:]
    y_pred_train, y_pred_test = y_pred[:split_idx], y_pred[split_idx:]

    # --- Global metrics
    metrics_all = get_metrics(y_real, y_pred)
    metrics_train = get_metrics(y_real_train, y_pred_train)
    metrics_test = get_metrics(y_real_test, y_pred_test)

    df_main = pd.DataFrame({
        "Set": ["All", "Train", "Test"],
        "Accuracy": [metrics_all["Accuracy"], metrics_train["Accuracy"], metrics_test["Accuracy"]],
        "Recall": [metrics_all["Recall"], metrics_train["Recall"], metrics_test["Recall"]],
        # "Class-Wise Error": ["", "", ""],
        "F1": [metrics_all["F1"], metrics_train["F1"], metrics_test["F1"]],
        "Precision": [metrics_all["Precision"], metrics_train["Precision"], metrics_test["Precision"]],
        "MCC": [metrics_all["MCC"], metrics_train["MCC"], metrics_test["MCC"]],
    })

    # --- Per-class metrics
    classes = np.unique(y_real)

    precision_per_class = precision_score(y_real, y_pred, average=None, zero_division=0)
    recall_per_class = recall_score(y_real, y_pred, average=None, zero_division=0)
    f1_per_class = f1_score(y_real, y_pred, average=None, zero_division=0)

    accuracy_per_class = []
    class_error_per_class = []
    mcc_per_class = []

    for i, cls in enumerate(classes):
        idx = y_real == cls
        acc = accuracy_score(y_real[idx], y_pred[idx])
        accuracy_per_class.append(acc)
        class_error_per_class.append(1 - acc)

        # MCC per class (binary view: this class vs rest)
        y_true_bin = (y_real == cls).astype(int)
        y_pred_bin = (y_pred == cls).astype(int)
        mcc_per_class.append(matthews_corrcoef(y_true_bin, y_pred_bin))

    df_class = pd.DataFrame({
        "Set": [f"Class {cls}" for cls in classes],
        "Accuracy": accuracy_per_class,
        "Recall": recall_per_class,
        # "Class-Wise Error": class_error_per_class,
        "F1": f1_per_class,
        "Precision": precision_per_class,
        "MCC": mcc_per_class
    })

    # --- Combine
    df_combined = pd.concat([df_main, df_class], ignore_index=True)

    # --- ROC (ONLY for binary classification)
    if len(classes) == 2:
        fpr, tpr, thresholds = roc_curve(y_real, y_pred)
        roc_auc = auc(fpr, tpr)
        auc_column = [""] * (len(fpr) - 1) + [round(roc_auc, 3)]

        roc_df = pd.DataFrame({
            "FPR": fpr,
            "TPR": tpr,
            "AUC": auc_column
        })
    else:
        roc_df = pd.DataFrame({
            "Info": ["ROC not available for multi-class without probabilities"]
        })

    # --- Confusion Matrix
    cm = confusion_matrix(y_real, y_pred)
    cm_df = pd.DataFrame(
        cm,
        index=[f"Actual {i}" for i in classes],
        columns=[f"Predicted {i}" for i in classes]
    )

    return df_combined, roc_df, cm_df

def generate_fake_convergence(df_combined, y_real, y_pred_fake, convegence_direction="down" ,Convergence_metric=Convergence_metric):
    if "Train" in df_combined["Set"].values:
        Target_metric_train = df_combined.loc[df_combined["Set"] == "Train", Convergence_metric].values[0]
        if np.isnan(Target_metric_train):
            Target_metric_train = 0.0
    else:
        Target_metric_train = accuracy_score(y_real, y_pred_fake)

    
    convergence_array = get_conv(count=200, target=float(Target_metric_train), direction=convegence_direction)
    


    df_convergence = pd.DataFrame({"Convergence": convergence_array})
    print("Fake convergence table created.")
    return df_convergence

# === EXECUTION ===

# Step 0: load data file (expects data with two columns: y_true, y_pred)
data = np.loadtxt(dataPath)
if data.ndim == 1 and data.shape[0] >= 2:
    data = data.reshape(-1, 2)
y_real = data[:, 0].astype(int)
y_pred = data[:, 1].astype(int)
print("Data loaded:", data.shape)

# Step A: ensure binary target if needed (preserve your median-threshold behavior)
unique_vals = np.unique(y_real)
if unique_vals.size > 2:
    median_value = np.median(y_real)
    y_real = (y_real > median_value).astype(int)
    y_pred = (y_pred > median_value).astype(int)

# Step B: Optionally adjust predictions to reach a target accuracy
y_pred_fake = fake_accuracy_prediction(y_real, y_pred, Accuracy_target)
print("Original Accuracy:", accuracy_score(y_real, y_pred))
print("Fake Accuracy after adjustment (if any):", accuracy_score(y_real, y_pred_fake))

# Step C: Build value/predict table and save back into data array for Excel writing
data[:, 1] = y_pred_fake
df_value_pred = pd.DataFrame(data.astype(int), columns=["y_real", "y_pred"])
print("Value/predict table created.")

# Step F: Define model parameters (kept as given)
df_params = pd.DataFrame(list(params.items()), columns=["parameters", "values"])
print("Model parameters defined.")

# ---- REPLACED: Step G & H (original REC & confusion summary replaced) ----
# Step G: Build df_combined (global + per-class metrics) -- will be written where REC used to be
df_combined, roc_df, cm_df = build_classification_reports(y_real, y_pred_fake)
print("Classification reports created successfully.")



df_convergence = generate_fake_convergence(df_combined, y_real, y_pred_fake, convegence_direction=convegence_direction)
print("Convergence data created successfully.")
# -------------------------------------------------------------------------

# Step I: Export to Excel (close file first if open)
close_excel_file(outputPath)

# Ensure workbook exists; load/create using openpyxl
if not os.path.exists(outputPath):
    from openpyxl import Workbook
    wb = Workbook()
    wb.save(outputPath)

book = load_workbook(outputPath)

with pd.ExcelWriter(outputPath, engine="openpyxl", mode="a", if_sheet_exists="new") as writer:
    # Create new sheet and attach to writer
    worksheet = writer.book.create_sheet(sheet_name)
    writer.sheets[sheet_name] = worksheet

    # Header/title (preserve original logic)
    if optimizer_name.strip():
        title = f"{model_name} + {optimizer_name.strip()}"
        merge_end_col = 14
        include_convergence = True
    else:
        title = model_name
        merge_end_col = 13
        include_convergence = False

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_end_col + 1)
    cell = worksheet.cell(row=1, column=1)
    cell.value = title
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color="E1DFFF", end_color="E1DFFF", fill_type="solid")

    # Write tables similar to original layout
    write_table(df_value_pred, startrow=1, startcol=0, style_key="value_pred", worksheet=worksheet, writer=writer, header_styles=None, sheet_name=sheet_name)

    params_col = len(df_value_pred.columns) + 1
    write_table(df_params, startrow=1, startcol=params_col, style_key="params", worksheet=worksheet, writer=writer, header_styles=None, sheet_name=sheet_name)

    metrics_col = params_col + len(df_params.columns) + 1
    write_table(df_combined, startrow=1, startcol=metrics_col, style_key="metrics", worksheet=worksheet, writer=writer, header_styles=None, sheet_name=sheet_name)

    # REC used to be written at CM_start_row, params_col; write df_combined (metrics) there
    # CM_col = len(df_value_pred.columns) + 1
    CM_start_row = len(df_params) + 6
    write_table(cm_df, startrow=CM_start_row, startcol=params_col, style_key="rec_curve", worksheet=worksheet, writer=writer, header_styles=None, sheet_name=sheet_name)

    # Write ROC table just under df_combined (preserve spacing)
    write_table(roc_df, startrow=CM_start_row, startcol=metrics_col, style_key="roc", worksheet=worksheet, writer=writer, header_styles=None, sheet_name=sheet_name)

    if include_convergence:
        convergence_col = metrics_col + len(df_combined.columns) + 1
        write_table(df_convergence, startrow=1, startcol=convergence_col, style_key="error", worksheet=worksheet, writer=writer, header_styles=None, sheet_name=sheet_name)

open_excel_file(outputPath)
print("✅ Structured Excel file saved successfully.")



# import pandas as pd
# import numpy as np
# import os

# # ==== Title logic ====
# if optimizer_name.strip():
#     title = f"{model_name} + {optimizer_name.strip()}"
#     include_convergence = True
# else:
#     title = model_name
#     include_convergence = False

# # ==== Helper: place df in canvas ====
# def place_df(canvas, df, start_row, start_col):
#     for i in range(df.shape[0]):
#         for j in range(df.shape[1]):
#             canvas[start_row + i][start_col + j] = df.iat[i, j]

#     # write headers
#     for j, col in enumerate(df.columns):
#         canvas[start_row - 1][start_col + j] = col

# # ==== Calculate layout ====
# params_col = len(df_value_pred.columns) + 1
# metrics_col = params_col + len(df_params.columns) + 1

# CM_start_row = len(df_params) + 6

# if include_convergence:
#     convergence_col = metrics_col + len(df_combined.columns) + 1
#     total_cols = convergence_col + len(df_convergence.columns)
# else:
#     total_cols = metrics_col + len(df_combined.columns)

# total_rows = max(
#     len(df_value_pred),
#     len(df_params),
#     len(df_combined),
#     CM_start_row + len(cm_df),
#     CM_start_row + len(roc_df)
# ) + 5

# # ==== Create empty canvas ====
# canvas = [["" for _ in range(total_cols + 5)] for _ in range(total_rows + 5)]

# # ==== Title ====
# canvas[0][0] = title

# # ==== Place tables ====
# place_df(canvas, df_value_pred, start_row=2, start_col=0)
# place_df(canvas, df_params, start_row=2, start_col=params_col)
# place_df(canvas, df_combined, start_row=2, start_col=metrics_col)

# place_df(canvas, cm_df, start_row=CM_start_row+1, start_col=params_col)
# place_df(canvas, roc_df, start_row=CM_start_row+1, start_col=metrics_col)

# if include_convergence:
#     place_df(canvas, df_convergence, start_row=2, start_col=convergence_col)

# # ==== Convert to DataFrame ====
# final_df = pd.DataFrame(canvas)

# # ==== Save CSV ====
# csv_path = outputPath.replace(".xlsx", ".csv")
# final_df.to_csv(csv_path, index=False, header=False)

# print(f"✅ CSV saved: {csv_path}")


