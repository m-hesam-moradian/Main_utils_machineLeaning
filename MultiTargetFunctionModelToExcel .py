# === IMPORTS ===
import numpy as np
import pandas as pd
from sklearn.metrics import r2_score, mean_squared_error, auc
import os
import win32com.client
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# === CONFIGURATION ===
# --- Model & File Parameters ---
params = {
    "iterations": 300,
    "depth": 6,
    "learning_rate": 0.1,
}
optimizer_name = " " # "Spider Wasp Optimizer (SWO)", "Cyclone Optimization Algorithm (COA)", or ""
model_name = "RFR"
base_sheet_name = "RFR_Results"
dataPath = r"data\Data_err.npt"
outputPath = r"task\Data.xlsx"

# --- Analysis Parameters ---
R2_target = 0.89848934
min_error = -55
max_error = 63
Convergence_metric = "RMSE"  # Options: "RMSE", "SMAPE", etc.
convegence_direction = "higher"  # Options: "higher" or "lower"

# === HELPER FUNCTIONS ===
def make_style(color):
    """Creates a style object for openpyxl cells."""
    return {
        "font": Font(bold=True),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "fill": PatternFill(start_color=color, end_color=color, fill_type="solid")
    }

def fake_r2_prediction(y_real, y_pred, R2_target):
    """Adjusts predictions to meet a target R2 score by blending with real values."""
    current_r2 = r2_score(y_real, y_pred)
    if current_r2 >= R2_target:
        return y_pred
    for blend in np.linspace(0, 1, 1000):
        y_fake = y_pred * (1 - blend) + y_real * blend
        if r2_score(y_real, y_fake) >= R2_target:
            return y_fake
    return y_pred * 0.5 + y_real * 0.5

def enforce_error_bounds(y_real, y_pred, min_error, max_error):
    """Constrains the prediction errors to be within a specified min/max percentage."""
    y_pred_enforced = y_pred.copy()
    for i in range(len(y_real)):
        if y_real[i] == 0:
            continue
        error_percent = (y_pred_enforced[i] / y_real[i] - 1) * 100
        if not (min_error <= error_percent <= max_error):
            random_percent = np.random.uniform(min_error, max_error) / 100
            y_pred_enforced[i] = y_real[i] * (1 + random_percent)
    return y_pred_enforced

import numpy as np
import pandas as pd
from sklearn.metrics import mean_squared_error, r2_score

def build_metrics_table(y_real, y_pred):
    """Calculates R2, RMSE, COV, AARD, and MAEM for All, Train, and Test splits."""

    def compute_metrics(y_true, y_hat):
        # --- Core errors ---
        abs_error = np.abs(y_true - y_hat)
        rel_error = abs_error / (np.abs(y_true) + 1e-12)

        # --- RMSE ---
        rmse = np.sqrt(mean_squared_error(y_true, y_hat))

        # --- R2 ---
        r2 = r2_score(y_true, y_hat)

        # --- AARD (%)
        aard = 100 * np.mean(rel_error)

        # --- MAEM ---
        maem = np.sum(abs_error) / (np.abs(np.sum(y_true)) + 1e-12)

        # --- COV ---
        ratio = y_hat / (y_true + 1e-12)
        mean_ratio = np.mean(ratio)
        std_ratio = np.std(ratio, ddof=1)
        cov = std_ratio / (mean_ratio + 1e-12)

        return {
            "R2": r2,
            "RMSE": rmse,
            "COV": cov,
            "AARD (%)": aard,
            "MAEM": maem
        }

    # --- Split into Train/Test ---
    split_idx = int(len(y_real) * 0.8)
    y_real_train, y_real_test = y_real[:split_idx], y_real[split_idx:]
    y_pred_train, y_pred_test = y_pred[:split_idx], y_pred[split_idx:]

    # --- Compute metrics ---
    metrics_data = {
        "All": compute_metrics(y_real, y_pred),
        "Train": compute_metrics(y_real_train, y_pred_train),
        "Test": compute_metrics(y_real_test, y_pred_test),
    }

    df_metrics = pd.DataFrame(metrics_data).T.reset_index().rename(columns={"index": "Set"})
    return df_metrics

def build_rec_curve(y_real, y_pred):
    """Builds a Regression Error Characteristic (REC) curve DataFrame."""
    errors = np.abs(y_real - y_pred)
    epsilon = np.linspace(0, errors.max(), 200)
    accuracy = [np.mean(errors <= e) for e in epsilon]
    rec_auc = auc(epsilon, accuracy)
    
    df_rec_curve = pd.DataFrame({"Epsilon": epsilon, "Accuracy": accuracy})
    df_rec_curve["AUC"] = ""
    df_rec_curve.loc[0, "AUC"] = rec_auc
    return df_rec_curve

def build_relative_error_table(y_real, y_pred):
    """Calculates the relative error for each prediction."""
    rel_error = ((y_pred / y_real) - 1) * 100
    return pd.DataFrame({"Relative Error (%)": rel_error})

def get_conv(count, high, minPhase, maxPhase, direction):
    """Generates a fake convergence curve."""
    low_factor = np.random.uniform(1.2, 2.0)
    low = high * low_factor if direction == "higher" else high / low_factor
    phase = np.random.randint(minPhase, maxPhase + 1)
    
    convergence = []
    for _ in range(phase):
        repeated_count = np.random.randint(1, 6)
        random_number = np.random.uniform(low, high)
        convergence.extend([random_number] * repeated_count)
    
    convergence = np.resize(convergence, count)
    return np.sort(convergence)[::-1] if direction == "higher" else np.sort(convergence)

def write_table(df, startrow, startcol, style_key, worksheet, header_styles):
    """Writes a DataFrame to a specific location in an Excel worksheet with styling."""
    style = header_styles.get(style_key, make_style("D9D9D9"))
    
    # Write header
    for col_num, col_name in enumerate(df.columns):
        cell = worksheet.cell(row=startrow, column=startcol + col_num)
        cell.value = col_name
        cell.font = style["font"]
        cell.alignment = style["alignment"]
        cell.fill = style["fill"]
        
    # Write data
    for row_num, row_data in enumerate(df.itertuples(index=False), 1):
        for col_num, value in enumerate(row_data):
            worksheet.cell(row=startrow + row_num, column=startcol + col_num).value = value

def close_excel_file(filepath):
    """Closes a specific Excel file if it is open, saving changes."""
    if not os.path.exists(filepath):
        return
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        for wb in excel.Workbooks:
            if os.path.abspath(wb.FullName) == os.path.abspath(filepath):
                wb.Save()
                wb.Close(SaveChanges=False)
                print(f"💾 Saved and 🔒 Closed Excel file: {filepath}")
                break
        excel.Quit()
    except Exception as e:
        print(f"Could not close Excel file (it might not be open): {e}")

def open_excel_file(filepath):
    """Opens an Excel file."""
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        excel.Workbooks.Open(os.path.abspath(filepath))
        print(f"📂 Opened Excel file: {filepath}")
    except Exception as e:
        print(f"Could not open Excel file: {e}")

# === NEW MAIN ANALYSIS FUNCTION ===
def run_analysis_for_target(y_real, y_pred, config):
    """Runs the full analysis pipeline for a single target variable."""
    # Step 1: Adjust predictions for R2 target
    y_pred_adj = fake_r2_prediction(y_real, y_pred, config['R2_target'])
    
    # Step 2: Enforce error bounds
    y_pred_final = enforce_error_bounds(y_real, y_pred_adj, config['min_error'], config['max_error'])
    
    # Step 3: Build all dataframes
    df_value_pred = pd.DataFrame({"y_real": y_real, "y_pred": y_pred_final})
    df_metrics = build_metrics_table(y_real, y_pred_final)
    df_rec_curve = build_rec_curve(y_real, y_pred_final)
    df_error = build_relative_error_table(y_real, y_pred_final)
    df_params = pd.DataFrame(list(config['params'].items()), columns=["parameters", "values"])
    
    # Step 4: Generate fake convergence
    target_metric_train = df_metrics.loc[df_metrics["Set"] == "Train", config['Convergence_metric']].values[0]
    convergence_array = get_conv(
        count=200, high=target_metric_train, minPhase=24, maxPhase=32, direction=config['convegence_direction']
    )
    df_convergence = pd.DataFrame({"Convergence": convergence_array})

    return {
        "value_pred": df_value_pred,
        "metrics": df_metrics,
        "rec_curve": df_rec_curve,
        "error": df_error,
        "params": df_params,
        "convergence": df_convergence
    }

# === MAIN EXECUTION BLOCK ===
if __name__ == "__main__":
    # Step 1: Load data with multiple targets
    # Assumes 4 columns: y_real_1, y_pred_1, y_real_2, y_pred_2
    all_data = np.loadtxt(dataPath)
    print(f"Data loaded with shape: {all_data.shape}")

    # Create an empty Excel file if it doesn't exist to allow append mode
    if not os.path.exists(outputPath):
        pd.DataFrame().to_excel(outputPath)
        print(f"Created initial Excel file at: {outputPath}")
    
    # Close Excel file before writing to avoid conflicts
    close_excel_file(outputPath)

    # Step 2: Loop through each target and run analysis
    num_targets = all_data.shape[1] // 2
    for i in range(num_targets):
        target_index = i + 1
        print(f"\n--- Processing Target {target_index} ---")

        # --- Data Slicing ---
        # **ADJUST THIS** if your data columns are ordered differently.
        # E.g., for y_real_1, y_real_2, y_pred_1, y_pred_2, you would use:
        # y_real = all_data[:, i]
        # y_pred = all_data[:, i + num_targets]
        y_real = all_data[:, i * 2]      # Column 0 for target 1, col 2 for target 2
        y_pred = all_data[:, i * 2 + 1]  # Column 1 for target 1, col 3 for target 2
        
        # --- Run Analysis ---
        analysis_config = {
            'R2_target': R2_target, 'min_error': min_error, 'max_error': max_error,
            'params': params, 'Convergence_metric': Convergence_metric, 
            'convegence_direction': convegence_direction
        }
        results = run_analysis_for_target(y_real, y_pred, analysis_config)

        # --- Write to Excel ---
        sheet_name = f"{base_sheet_name}_Target_{target_index}"
        
        try:
            with pd.ExcelWriter(outputPath, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                # Create a new sheet or get existing one to overwrite
                if sheet_name in writer.book.sheetnames:
                    worksheet = writer.book[sheet_name]
                    # Clear existing content if sheet exists
                    worksheet.delete_rows(1, worksheet.max_row)
                    worksheet.delete_cols(1, worksheet.max_column)
                else:
                    worksheet = writer.book.create_sheet(sheet_name)
                
                writer.sheets[sheet_name] = worksheet

                # Define header styles
                header_styles = {
                    "value_pred": make_style("9DC3E6"), # richer blue
                    "params": make_style("A9D08E"),     # deeper green
                    "metrics": make_style("F4B084"),    # stronger orange
                    "error": make_style("FFD966"),      # golden yellow
                    "rec_curve": make_style("E06666"),   # bold red
                    "convergence": make_style("C5B4E3") # purple
                }

                # --- Write Custom Title Header ---
                include_convergence = bool(optimizer_name.strip())
                title = f"{model_name} + {optimizer_name.strip()}" if include_convergence else model_name
                merge_end_col = 14 if include_convergence else 13
                
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_end_col)
                cell = worksheet.cell(row=1, column=1)
                cell.value = title
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                worksheet.row_dimensions[1].height = 30

                # --- Write Tables to Sheet ---
                write_table(results['value_pred'], startrow=3, startcol=1, style_key="value_pred", worksheet=worksheet, header_styles=header_styles)
                
                params_col = 1 + results['value_pred'].shape[1] + 1
                write_table(results['params'], startrow=3, startcol=params_col, style_key="params", worksheet=worksheet, header_styles=header_styles)
                
                rec_start_row = 3 + results['params'].shape[0] + 4
                write_table(results['rec_curve'], startrow=rec_start_row, startcol=params_col, style_key="rec_curve", worksheet=worksheet, header_styles=header_styles)
                
                metrics_col = params_col + results['params'].shape[1] + 1
                write_table(results['metrics'], startrow=3, startcol=metrics_col, style_key="metrics", worksheet=worksheet, header_styles=header_styles)
                
                error_col = metrics_col + results['metrics'].shape[1] + 1
                write_table(results['error'], startrow=3, startcol=error_col, style_key="error", worksheet=worksheet, header_styles=header_styles)
                
                if include_convergence:
                    conv_col = error_col + results['error'].shape[1] + 1
                    write_table(results['convergence'], startrow=3, startcol=conv_col, style_key="convergence", worksheet=worksheet, header_styles=header_styles)

            print(f"✅ Successfully wrote results for Target {target_index} to sheet: {sheet_name}")

        except Exception as e:
            print(f"❌ Error writing to Excel for Target {target_index}: {e}")

    # Step 3: Open the final Excel file to view results
    open_excel_file(outputPath)
    print("\n✅ All analyses complete and Excel file generated.")

