import pandas as pd
import shap
from sklearn.model_selection import train_test_split
from lightgbm import LGBMClassifier


def shap_analysis(
    model,
    X_train,
    y_train,
    X_test,
    y_test,
    save_path=None,
    sheet_name="SHAP_Sensitivity",
):
    """
    Perform SHAP analysis for any trained model.

    Parameters:
        model: A trained ML model (e.g., HistGradientBoostingClassifier).
        X_train: Training features (DataFrame or array).
        y_train: Training target (Series or array).
        X_test: Test features (DataFrame or array).
        y_test: Test target (Series or array).
        save_path: Optional, path to Excel file to save SHAP sensitivity results.
        sheet_name: Name of the Excel sheet if saving results.

    Returns:
        sensitivity_df: DataFrame containing feature sensitivity scores.
        shap_values: The SHAP values object (for plots or further analysis).
    """
    if not hasattr(model, "fit"):
        raise ValueError("Provided model is not a valid scikit-learn compatible model.")

    try:
        model.predict(X_train)
    except:
        model.fit(X_train, y_train)

    explainer = shap.Explainer(model, X_train, feature_names=X_train.columns)
    shap_values = explainer.shap_values(X_test)

    feature_names = (
        X_test.columns
        if hasattr(X_test, "columns")
        else [f"Feature_{i}" for i in range(X_test.shape[1])]
    )
    shap_df = pd.DataFrame(shap_values, columns=feature_names)
    shap_df["BaseValue"] = explainer.expected_value
    shap_df["ModelPrediction"] = (
        shap_df[feature_names].sum(axis=1) + shap_df["BaseValue"]
    )

    sensitivity_df = (
        pd.DataFrame({
            "Feature": feature_names,
            "Sensitivity": shap_df[feature_names].abs().mean(),
        })
        .sort_values(by="Sensitivity", ascending=False)
        .reset_index(drop=True)
    )

    if save_path:
        from openpyxl import load_workbook

        try:
            book = load_workbook(save_path)
            if sheet_name in book.sheetnames:
                book.remove(book[sheet_name])
                book.save(save_path)
        except FileNotFoundError:
            pass
        with pd.ExcelWriter(save_path, engine="openpyxl", mode="a") as writer:
            sensitivity_df.to_excel(writer, sheet_name=sheet_name, index=False)

    return sensitivity_df, shap_values

# --- Load dataset ---
sheet_name = "Data_after_KFold_LGBC"
file_path = r"C:\Users\Sam\Desktop\ML\task\Data.xlsx"

df = pd.read_excel(file_path, sheet_name=sheet_name).dropna()
target_column = df.columns[-1]

# --- Features and Target ---
X = df.drop(columns=[target_column])
y = df[target_column]

# --- Train-Test Split ---
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, shuffle=False)

# --- Run SHAP Analysis with ADAC ---
sensitivity_df_shap, shap_values = shap_analysis(
    model=LGBMClassifier(),
    X_train=X_train,
    y_train=y_train,
    X_test=X_test,
    y_test=y_test,
)

# --- Export to clipboard ---
sensitivity_df_shap.to_clipboard(index=False)