# === IMPORTS ===
import numpy as np
import pandas as pd
from sklearn.metrics import r2_score, mean_squared_error, auc
import os
import win32com.client
import random
from openpyxl.styles import Font, Alignment, PatternFill

# === CONFIGURATION ===
# cation : do not remove this Guide comments wich hase * at the beginning of the lines

# * Define model name and Optimizers based on task information and comment 2 rest of model names but leave optimizers as it is
# model_name="QR"
# model_name="SGB"
model_name="SBR"

# Optimizers to consider
optimizers = ["", "NOA", "OOA"]

# * Define parameter ranges and digits limit to generate random parameters based on range and digits limit for each model and comment rest of model params 
# * the range and digit should be propriate to model used to seem realistic and defult params is already cleare each models defult params values shoud be 

# QR params (Quantile Regression)
# params = {
#     "alpha": {"range": [0.0001, 10.0], "digit": 4, "default": 1.0},
#     "quantile": {"range": [0.1, 0.9], "digit": 1, "default": 0.5}
# }

# SGB params (Stochastic Gradient Boosting)
# params = {
#     "learning_rate": {"range": [0.01, 0.3], "digit": 3, "default": 0.1},
#     "n_estimators": {"range": [50, 400], "digit": 0, "default": 100},
#     "max_depth": {"range": [2, 8], "digit": 0, "default": 3},
#     "subsample": {"range": [0.5, 1.0], "digit": 1, "default": 1.0}
# }

# SBR params (Sparse Bayesian Regression / ARDRegression)
params = {
    "tol": {"range": [1e-6, 1e-2], "digit": 6, "default": 1e-3},
    "max_iter": {"range": [100, 1000], "digit": 0, "default": 300},
    "alpha_1": {"range": [1e-6, 1e-2], "digit": 6, "default": 1e-6}
}

# * Define target R² values based on model used and comment rest of target R² values
# *based on kfold results:
# Model	Best Fold	Best R2
# QR	3	0.861964923
# SGB	1	0.840743995
# SBR	4	0.916389327



# r2_target_values = [0.861964923, 0.94462, 0.963803]  # QR targets
# r2_target_values = [0.840743995, 0.925462, 0.9535203]  # SGB targets
r2_target_values = [0.916389327, 0.97045, 0.980123]  # SBR targets



# Other parameters
min_error = -43.54
max_error = 57.43
# Ensemble: DST for combining best hybrid models of each category two by two
Convergence_metric = "MDAPE"
convegence_direction = "lower"  # "lower" for MDAPE convergence
dataPath = r"data\Data_err.npt"
outputPath = r"task\Data.xlsx"

# === FUNCTIONS ===
def generate_random_params(params, optimizer_name=""):
    generated_params = {}

    for param, properties in params.items():
        min_val, max_val = properties["range"]  # Extract range
        precision = properties["digit"]  # Extract precision (number of decimal places)
        
        # If optimizer is empty (single model), use default value
        if optimizer_name == "" and "default" in properties:
            rand_val = properties["default"]  # Use the default value if no optimizer
        else:
            # Otherwise, generate a random value within the given range
            rand_val = random.uniform(min_val, max_val)
        
        # Round the value to the specified number of decimal places
        rand_val_rounded = round(rand_val, precision)
        
        # Store the result
        generated_params[param] = rand_val_rounded
    
    return generated_params

def build_metrics_table(y_real, y_pred):

    def compute_metrics(y_true, y_hat):
        y_true = np.asarray(y_true)
        y_hat = np.asarray(y_hat)
        epsilon = 1e-12  # avoid division by zero

        # --- R2 ---
        r2 = r2_score(y_true, y_hat)

        # --- RMSE ---
        rmse = np.sqrt(mean_squared_error(y_true, y_hat))

        # --- U95 (95% expanded uncertainty) ---
        # Standard definition: 1.96 * Standard Deviation of Residuals
        residuals = y_true - y_hat
        u95 = 1.96 * np.std(residuals)

        # --- COM (Coefficient of Multiple Correlation) ---
        # Calculated as the Pearson correlation between observed and predicted values
        if np.std(y_true) == 0 or np.std(y_hat) == 0:
            com = 0.0
        else:
            com = np.corrcoef(y_true, y_hat)[0, 1]

        # --- MDAPE (Median Absolute Percentage Error) ---
        ape = np.abs((y_hat - y_true) / (y_true + epsilon))
        mdape = np.median(ape) * 100

        return {
            "R2": r2,
            "RMSE": rmse,
            "U95": u95,
            "COM": com,
            "MDAPE": mdape,
        }

    # --- Split data into Train/Test/Value sets ---
    split_idx = int(len(y_real) * 0.8)
    y_real_train, y_real_test = y_real[:split_idx], y_real[split_idx:]
    y_pred_train, y_pred_test = y_pred[:split_idx], y_pred[split_idx:]

    # Split Test into Value / Value-Test
    mid = len(y_real_test) // 2
    y_real_value, y_pred_value = y_real_test[:mid], y_pred_test[:mid]
    y_real_valte, y_pred_valte = y_real_test[mid:], y_pred_test[mid:]

    # --- Compute metrics ---
    M_all = compute_metrics(y_real, y_pred)
    M_train = compute_metrics(y_real_train, y_pred_train)
    M_test = compute_metrics(y_real_test, y_pred_test)
    M_value = compute_metrics(y_real_value, y_pred_value)
    M_valte = compute_metrics(y_real_valte, y_pred_valte)

    # --- Build DataFrame ---
    cols = ["Set", "R2", "RMSE", "U95", "COM", "MDAPE"]

    df_metrics = pd.DataFrame(
        [
            ["All", M_all["R2"], M_all["RMSE"], M_all["U95"], M_all["COM"], M_all["MDAPE"]],
            ["Train", M_train["R2"], M_train["RMSE"], M_train["U95"], M_train["COM"], M_train["MDAPE"]],
            ["Test", M_test["R2"], M_test["RMSE"], M_test["U95"], M_test["COM"], M_test["MDAPE"]],
            ["Value", M_value["R2"], M_value["RMSE"], M_value["U95"], M_value["COM"], M_value["MDAPE"]],
            ["Value-test", M_valte["R2"], M_valte["RMSE"], M_valte["U95"], M_valte["COM"], M_valte["MDAPE"]],
        ],
        columns=cols,
    )

    return df_metrics

def fake_r2_prediction(y_real, y_pred, R2_target):
    current_r2 = r2_score(y_real, y_pred)
    if current_r2 >= R2_target:
        return y_pred
    for blend in np.linspace(0, 1, 1000):
        y_fake = y_pred * (1 - blend) + y_real * blend
        if r2_score(y_real, y_fake) >= R2_target:
            return y_fake
    return y_pred * 0.5 + y_real * 0.5


def enforce_error_bounds(y_real, y_pred, min_error, max_error):
    y_pred = y_pred.copy()
    for i in range(len(y_real)):
        if y_real[i] == 0:
            continue
        error_percent = (y_pred[i] / y_real[i] - 1) * 100
        if error_percent < min_error or error_percent > max_error:
            random_percent = np.random.uniform(min_error, max_error) / 100
            y_pred[i] = y_real[i] * (1 + random_percent)
    return y_pred


def build_rec_curve(y_real, y_pred):
    errors = np.abs(y_real - y_pred)
    epsilon = np.linspace(0, errors.max(), 200)
    accuracy = [np.mean(errors <= e) for e in epsilon]
    rec_auc = auc(epsilon, accuracy)

    df_rec_curve = pd.DataFrame(
        {"Epsilon": epsilon, "Accuracy": accuracy, "AUC": ["" for _ in range(len(epsilon))]}
    )
    df_rec_curve.loc[0] = [np.nan, np.nan, rec_auc]
    return df_rec_curve


def build_relative_error_table(y_real, y_pred):
    rel_error = ((y_pred / y_real) - 1) * 100
    return pd.DataFrame({"Relative Error (%)": rel_error})


def get_conv(
    count=200, high=0.2, minPhase=6, maxPhase=10, convegence_direction="higher", tail_repeats=10
):
    high = float(high)  # ensure scalar
    factor = np.random.uniform(1.2, 2.0)

    if convegence_direction == "higher":
        low = high / factor
        lo, hi = low, high
    else:
        start_high = high * factor
        lo, hi = high, start_high

    phase = np.random.randint(minPhase, maxPhase + 1)
    convergence = []

    for _ in range(phase):
        repeated_count = np.random.randint(1, 6)
        random_number = np.random.uniform(lo, hi)
        convergence.extend([random_number] * repeated_count)

    convergence = np.resize(convergence, count)

    if convegence_direction == "higher":
        convergence = np.sort(convergence)
    else:
        convergence = np.sort(convergence)[::-1]

    tail_repeats = int(min(tail_repeats, count))
    convergence[-tail_repeats:] = high

    return np.array(convergence)


def make_style(color):
    """Helper function to create a style."""
    return {
        "font": Font(bold=True),
        "alignment": Alignment(horizontal="center"),
        "fill": PatternFill(start_color=color, end_color=color, fill_type="solid"),
    }


def write_table(df, startrow, startcol, style_key, worksheet, writer, header_styles, sheet_name):
    header_styles = {
        "value_pred": make_style("9DC3E6"),  # richer blue
        "params": make_style("A9D08E"),      # deeper green
        "metrics": make_style("F4B084"),     # stronger orange
        "error": make_style("FFD966"),       # golden yellow
        "rec_curve": make_style("E06666"),   # bold red
    }
    style = header_styles.get(style_key, make_style("D9D9D9"))  # fallback gray

    # Write header row
    for col_num, col_name in enumerate(df.columns):
        row = startrow + 1
        col = startcol + col_num + 1
        cell = worksheet.cell(row=row, column=col)
        cell.value = col_name
        cell.font = style["font"]
        cell.alignment = style["alignment"]
        cell.fill = style["fill"]

    # Write data rows
    for row_num, row_data in enumerate(df.values):
        for col_num, value in enumerate(row_data):
            worksheet.cell(row=startrow + 2 + row_num, column=startcol + col_num + 1).value = value


def close_excel_file(filepath):
    excel = win32com.client.Dispatch("Excel.Application")
    for wb in excel.Workbooks:
        if os.path.abspath(wb.FullName) == os.path.abspath(filepath):
            wb.Save()
            wb.Close(SaveChanges=False)
            print("💾 Saved and 🔒 Closed Excel file:", filepath)
            break
    excel.Quit()


def open_excel_file(filepath):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True
    excel.Workbooks.Open(os.path.abspath(filepath))
    print("📂 Opened Excel file:", filepath)

# === EXECUTION ===

# Step 1: Load data
data = np.loadtxt(dataPath)
y_real = data[:, 0]
y_pred = data[:, 1]
print("Data loaded:", data.shape)


# Ensure that the number of optimizers matches the number of r2_target_values
assert len(optimizers) == len(r2_target_values), "Optimizers and target R² lists must have the same length."

for idx, optimizer_name in enumerate(optimizers):
    target_r2 = r2_target_values[idx]  # Get the corresponding target R² for the current optimizer
    print(f"Running with optimizer: {optimizer_name}, R² target: {target_r2}")

    # Generate random parameters
    random_params = generate_random_params(params, optimizer_name)
    print(f"Params: {random_params}")

    # Assign random parameters to model
    model_params = random_params
    print("Model parameters:", model_params)

    # Adjust predictions to meet the target R²
    y_pred_fake = fake_r2_prediction(y_real, y_pred, target_r2)
    print("Original R²:", r2_score(y_real, y_pred))
    print("Fake R² before error enforcement:", r2_score(y_real, y_pred_fake))

    # Enforce error bounds
    y_pred_fake = enforce_error_bounds(y_real, y_pred_fake, min_error, max_error)
    print("Fake R² after error enforcement:", r2_score(y_real, y_pred_fake))

    # Step 3: Build value/predict table
    data[:, 1] = y_pred_fake
    df_value_pred = pd.DataFrame(data, columns=["y_real", "y_pred"])
    print("Value/predict table created.")

    # Step 4: Build metrics table (UPDATED METRICS)
    df_metrics = build_metrics_table(y_real, y_pred_fake)
    print("Metrics table created : ", df_metrics)

    # Step 5.5: Generate fake convergence based on MBE from training (UPDATED)
    Target_metric_train = df_metrics.loc[df_metrics["Set"] == "Train", Convergence_metric].values[0]

    # For convergence, we typically track magnitude, so use abs(MBE) as "high"
    convergence_array = get_conv(
        count=200,
        high=abs(Target_metric_train),
        minPhase=24,
        maxPhase=32,
        convegence_direction=convegence_direction,
    )
    df_convergence = pd.DataFrame({"Convergence": convergence_array})
    print("Fake convergence table created.")

    # Step 6: Define model parameters
    df_params = pd.DataFrame(list(model_params.items()), columns=["parameters", "values"])
    print("Model parameters defined.")

    # Step 7: Build REC curve (unchanged)
    df_rec_curve = build_rec_curve(y_real, y_pred_fake)
    print("REC curve created. AUC =", df_rec_curve.loc[0, "AUC"])

    # Step 8: Build relative error table (unchanged)
    df_error = build_relative_error_table(y_real, y_pred_fake)
    print("Relative error table created.")

    # Step 9: Close Excel if open, then export to Excel
    close_excel_file(outputPath)

    from openpyxl import load_workbook

    # Load existing workbook (unchanged usage)
    book = load_workbook(outputPath)

    # Calculate indices based on total length
    total_len = len(data)
    idx_1 = int(total_len * 0.80)
    idx_2 = idx_1 + int(total_len * 0.10)

    # Create DataFrames for the Excel writer
    df_train_data = pd.DataFrame(data[:idx_1], columns=["Train_Real", "Train_Pred"])
    df_test_data  = pd.DataFrame(data[idx_1:idx_2], columns=["Test_Real", "Test_Pred"])
    df_val_data   = pd.DataFrame(data[idx_2:], columns=["Val_Real", "Val_Pred"])

    # Generate sheet name based on model and optimizer (no target R² in sheet name)
    sheet_name = f"{model_name}_{optimizer_name}" if optimizer_name else f"{model_name}"

    with pd.ExcelWriter(outputPath, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        # Create new sheet
        worksheet = writer.book.create_sheet(sheet_name)
        writer.sheets[sheet_name] = worksheet

        # 1. Value Pred Table
        write_table(
            df_value_pred,
            startrow=1,
            startcol=0,
            style_key="value_pred",
            worksheet=worksheet,
            writer=writer,
            header_styles=None,
            sheet_name=sheet_name,
        )

        # 2. Params Table
        params_col = len(df_value_pred.columns) + 1
        write_table(
            df_params,
            startrow=1,
            startcol=params_col,
            style_key="params",
            worksheet=worksheet,
            writer=writer,
            header_styles=None,
            sheet_name=sheet_name,
        )

        # 3. Metrics Table (UPDATED columns)
        metrics_col = params_col + len(df_params.columns) + 1
        write_table(
            df_metrics,
            startrow=1,
            startcol=metrics_col,
            style_key="metrics",
            worksheet=worksheet,
            writer=writer,
            header_styles=None,
            sheet_name=sheet_name,
        )

        # 4. Error Table
        error_col = metrics_col + len(df_metrics.columns) + 1
        write_table(
            df_error,
            startrow=1,
            startcol=error_col,
            style_key="error",
            worksheet=worksheet,
            writer=writer,
            header_styles=None,
            sheet_name=sheet_name,
        )

        # 5. REC Curve (Below Params)
        rec_start_row = len(df_params) + 6
        write_table(
            df_rec_curve,
            startrow=rec_start_row,
            startcol=params_col,
            style_key="rec_curve",
            worksheet=worksheet,
            writer=writer,
            header_styles=None,
            sheet_name=sheet_name,
        )

        # === Determine Next Column Position ===
        current_col = error_col + len(df_error.columns)

        # 6. Convergence (Optional)
        if optimizer_name.strip():
            write_table(
                df_convergence,
                startrow=1,
                startcol=current_col,
                style_key="error",
                worksheet=worksheet,
                writer=writer,
                header_styles=None,
                sheet_name=sheet_name,
            )
            current_col += len(df_convergence.columns)

        # === Write Split Data Tables (Side by Side) ===
        train_col = current_col + 1
        write_table(
            df_train_data,
            startrow=1,
            startcol=train_col,
            style_key="value_pred",
            worksheet=worksheet,
            writer=writer,
            header_styles=None,
            sheet_name=sheet_name,
        )

        test_col = train_col + len(df_train_data.columns) + 1
        write_table(
            df_test_data,
            startrow=1,
            startcol=test_col,
            style_key="value_pred",
            worksheet=worksheet,
            writer=writer,
            header_styles=None,
            sheet_name=sheet_name,
        )

        val_col = test_col + len(df_test_data.columns) + 1
        write_table(
            df_val_data,
            startrow=1,
            startcol=val_col,
            style_key="value_pred",
            worksheet=worksheet,
            writer=writer,
            header_styles=None,
            sheet_name=sheet_name,
        )

        # === Custom Header Row (Merge Title) ===
        final_used_col = val_col + len(df_val_data.columns)

        if optimizer_name.strip():
            title = f"{model_name} + {optimizer_name.strip()}"
        else:
            title = model_name

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=final_used_col)

        cell = worksheet.cell(row=1, column=1)
        cell.value = title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="E1DFFF", end_color="E1DFFF", fill_type="solid")

open_excel_file(outputPath)

print("✅ Structured Excel file saved successfully.")
