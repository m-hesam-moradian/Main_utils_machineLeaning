# D:\ML\Project-2\src\model\train_model.py

import pandas as pd
from sklearn.model_selection import KFold
from sklearn.metrics import mean_squared_error, r2_score

from openpyxl import load_workbook


def get_best_fold_data_from_kf(X, y, scores, kf):
    """
    Uses provided KFold instance to return the train/test data
    for the fold with the highest R2 score.
    """
    best_fold_index = max(range(len(scores)), key=lambda i: scores[i]["R2"])
    print(
        f"Best fold index: {best_fold_index}, R2: {scores[best_fold_index]['R2']:.4f}"
    )
    for i, (train_idx, test_idx) in enumerate(kf.split(X)):
        if i == best_fold_index:
            X_train_best = X.iloc[train_idx].copy()
            X_test_best = X.iloc[test_idx].copy()
            y_train_best = y.iloc[train_idx].copy()
            y_test_best = y.iloc[test_idx].copy()
            return X_train_best, X_test_best, y_train_best, y_test_best

    raise ValueError("Best fold index not found in KFold split.")


def K_Fold_metrics(y_true, y_pred):
    metrics = {
        "R2": r2_score(y_true, y_pred),
        "RMSE": mean_squared_error(y_true, y_pred),
    }
    return metrics


def K_Fold(X, y, n_splits=5, DATA_PATH="", save_to_excel=False, model=0):
    # ✅ Only train model model
    kf = KFold(n_splits=n_splits, shuffle=False)
    K_Fold_Cross_Validation_Scores = []
    fold = 1
    for train_idx, val_idx in kf.split(X):
        print(f"\n🔁 Fold {fold} ------------------")
        X_train, X_val = X.iloc[train_idx], X.iloc[val_idx]
        y_train, y_val = y.iloc[train_idx], y.iloc[val_idx]

        # ✅ model Regressor

        model.fit(X_train, y_train)
        preds = model.predict(X_val)
        metrics = K_Fold_metrics(y_val, preds)
        K_Fold_Cross_Validation_Scores.append(metrics)

        print(f"  model → R2: {metrics['R2']:.4f}, RMSE: {metrics['RMSE']:.4f}")

        fold += 1
    print("\n✅ K-Fold Cross-Validation completed.")
    # ✅ Get best fold based on model score
    X_train_best, X_test_best, y_train_best, y_test_best = get_best_fold_data_from_kf(
        X, y, K_Fold_Cross_Validation_Scores, kf
    )
    # Combine features and target
    train_df = pd.concat([X_train_best.copy(), y_train_best.copy()], axis=1)
    test_df = pd.concat([X_test_best.copy(), y_test_best.copy()], axis=1)

    # Merge train and test vertically
    combined_df = pd.concat([train_df, test_df], axis=0).reset_index(drop=True)

    print("✅ Combined DataFrame using original target column name:")
    # Save combined K-Fold data to Excel

    if save_to_excel:
        book = load_workbook(DATA_PATH)
        if "DATA after K-Fold" in book.sheetnames:
            book.remove(book["DATA after K-Fold"])
            book.save(DATA_PATH)
        with pd.ExcelWriter(DATA_PATH, engine="openpyxl", mode="a") as writer:
            combined_df.to_excel(writer, sheet_name="DATA after K-Fold", index=False)

        K_Fold_Cross_Validation_Scores = pd.DataFrame(K_Fold_Cross_Validation_Scores)
        if "K-Fold" in book.sheetnames:
            book.remove(book["K-Fold"])
            book.save(DATA_PATH)
        with pd.ExcelWriter(DATA_PATH, engine="openpyxl", mode="a") as writer:
            K_Fold_Cross_Validation_Scores.to_excel(
                writer, sheet_name="K-Fold", index=False
            )
        print("✅ K_FOLD saved on excel")

    return (
        X_train_best,
        X_test_best,
        y_train_best,
        y_test_best,
        K_Fold_Cross_Validation_Scores,
        combined_df,
    )
