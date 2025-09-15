import os
import pandas as pd
from openpyxl import load_workbook


def csv_to_excel(
    csv_path: str, excel_path: str, sheet_name: str = "Sheet1", header: bool = True
):
    """
    Writes CSV data into an Excel file.
    - If Excel file exists, data is appended to the selected sheet.
    - If Excel file does not exist, a new file is created.

    Parameters
    ----------
    csv_path : str
        Path to the CSV file.
    excel_path : str
        Path to the Excel file (will be created if not exists).
    sheet_name : str, optional
        Sheet name to use (default: 'Sheet1').
    header : bool, optional
        Whether to include headers from the CSV (default: True).
    """

    # Load CSV into DataFrame
    df = pd.read_csv(csv_path)

    if not os.path.exists(excel_path):
        # ✅ Create a new Excel file
        df.to_excel(excel_path, sheet_name=sheet_name, index=False, header=header)
        print(f"📁 Created new Excel file: '{excel_path}' with sheet '{sheet_name}'.")
    else:
        # ✅ Append to existing Excel
        book = load_workbook(excel_path)

        if sheet_name not in book.sheetnames:
            # Create new sheet if not exists
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=header)
            print(
                f"➕ Created new sheet '{sheet_name}' in existing file '{excel_path}'."
            )
        else:
            # Append data
            sheet = book[sheet_name]
            start_row = sheet.max_row  # append after last row

            with pd.ExcelWriter(
                excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay"
            ) as writer:
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=start_row,
                    index=False,
                    header=header,
                )
            print(
                f"✅ Data from '{csv_path}' appended to '{excel_path}' in sheet '{sheet_name}'."
            )
