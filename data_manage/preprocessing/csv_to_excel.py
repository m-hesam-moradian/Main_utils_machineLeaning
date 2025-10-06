import os
import pandas as pd
from openpyxl import load_workbook


def csv_to_excel(csv_path, excel_path, sheet_name="Sheet1", header=True):
    df = pd.read_csv(csv_path)

    if not os.path.exists(excel_path):
        df.to_excel(excel_path, sheet_name=sheet_name, index=False, header=header)
        print(f"📁 Created new Excel file: '{excel_path}' with sheet '{sheet_name}'.")
        return

    book = load_workbook(excel_path)
    mode = "a"
    writer_args = dict(engine="openpyxl", mode=mode)

    if sheet_name not in book.sheetnames:
        with pd.ExcelWriter(excel_path, **writer_args) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=header)
        print(f"➕ Created new sheet '{sheet_name}' in existing file '{excel_path}'.")
    else:
        start_row = book[sheet_name].max_row
        writer_args["if_sheet_exists"] = "overlay"
        with pd.ExcelWriter(excel_path, **writer_args) as writer:
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=start_row,
                index=False,
                header=header,
            )
        print(
            f"✅ Appended data from '{csv_path}' to '{excel_path}' in sheet '{sheet_name}'."
        )


csv_to_excel(
    r"D:\ML\Main_utils\task\EI_No_3__Optimal Scheduling_Classification_DTC_RFR_XGBC_HOA_DOA_Data.csv",
    r"D:\ML\Main_utils\task\EI_No_3__Optimal Scheduling_Classification_DTC_RFR_XGBC_HOA_DOA_Data.xlsx",
    sheet_name="Logs",
)
