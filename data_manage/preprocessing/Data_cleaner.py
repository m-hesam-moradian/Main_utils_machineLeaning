import pandas as pd
from openpyxl import load_workbook



def close_excel_file(filepath):
    import os
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    for wb in excel.Workbooks:
        try:
            if os.path.abspath(wb.FullName) == os.path.abspath(filepath):
                wb.Save()
                wb.Close(SaveChanges=False)
                print("💾 Saved and 🔒 Closed Excel file:", filepath)
                break
        except Exception:
            pass
    excel.Quit()

def open_excel_file(filepath):
    import os
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True
    excel.Workbooks.Open(os.path.abspath(filepath))
    print("📂 Opened Excel file:", filepath)


def clean_missing_samples(excel_path, source_sheet="DATA", target_sheet="CLEANED_DATA"):

    close_excel_file(excel_path)
    # Load the data from the source sheet
    df = pd.read_excel(excel_path, sheet_name=source_sheet)

    # Drop rows with any missing values
    cleaned_df = df.dropna()

    # Load the workbook to check existing sheets
    book = load_workbook(excel_path)
    writer_args = dict(engine="openpyxl", mode="a")

    # If the target sheet already exists, remove it first
    if target_sheet in book.sheetnames:
        del book[target_sheet]
        book.save(excel_path)

    # Write the cleaned data to the new sheet
    with pd.ExcelWriter(excel_path, **writer_args) as writer:
        cleaned_df.to_excel(writer, sheet_name=target_sheet, index=False)
    print(f"🧹 Cleaned data saved to sheet '{target_sheet}' in '{excel_path}'.")

    cleaned_df.to_clipboard(index=False)
    open_excel_file(excel_path)
# Example usage
clean_missing_samples(
    excel_path=r"C:\Users\Sam\Desktop\ML\task\Data.xlsx",
    source_sheet="Data",
    target_sheet="Delete_missing_samples",
)
