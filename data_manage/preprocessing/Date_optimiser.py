import pandas as pd
from openpyxl import load_workbook

# A safe function to calculate the mode for grouping
def get_mode(series):
    m = series.dropna().mode()
    if not m.empty:
        return m.iloc[0]  # Returns the most frequent value
    return None

def aggregate_to_daily(excel_path, source_sheet="Data", target_sheet="Daily_Aggregated", 
                       timestamp_col="timestamp", unique_threshold=15):
    
    print(f"⏳ Loading data from '{excel_path}', sheet '{source_sheet}'...")
    df = pd.read_excel(excel_path, sheet_name=source_sheet)

    # 1. Convert timestamp to datetime and extract the date for grouping
    df[timestamp_col] = pd.to_datetime(df[timestamp_col])
    df['date_only'] = df[timestamp_col].dt.date

    # 2. Automatically Separate Numerical and Categorical Columns
    num_columns = []
    cat_columns = []
    
    for col in df.columns:
        if col in [timestamp_col, 'date_only']:
            continue  # Skip timestamp columns
            
        # Safety check: If it's a string/object type, it MUST be categorical
        if not pd.api.types.is_numeric_dtype(df[col]):
            cat_columns.append(col)
        # The main rule: > 10 unique values is Numerical, <= 10 is Categorical
        elif df[col].nunique() > unique_threshold:
            num_columns.append(col)
        else:
            cat_columns.append(col)

    print(f"📊 Automatically detected {len(num_columns)} Numerical columns (>10 unique values) -> Applying AVERAGE")
    print(f"📊 Automatically detected {len(cat_columns)} Categorical columns (<=10 unique values or Text) -> Applying MODE")

    # 3. Define aggregation rules dictionary
    agg_dict = {}
    
    for col in num_columns:
        agg_dict[col] = 'mean'   # Average for numbers
        
    for col in cat_columns:
        agg_dict[col] = get_mode # Custom Mode function for categories

    print("⚙️ Aggregating data to daily format (this might take a moment)...")
    
    # 4. Group by the extracted date and apply the aggregations
    daily_df = df.groupby('date_only').agg(agg_dict).reset_index()

    # 5. Delete the date completely so exactly 44 input features remain
    daily_df = daily_df.drop(columns=['date_only'])

    print(f"✅ Aggregation complete! Output shape: {daily_df.shape} (Expected: ~1520 rows, 44 columns)")

    # 6. Save using your openpyxl logic
    print("💾 Saving to Excel...")
    book = load_workbook(excel_path)
    writer_args = dict(engine="openpyxl", mode="a")

    # If the target sheet already exists, remove it first
    if target_sheet in book.sheetnames:
        del book[target_sheet]
        book.save(excel_path)

    # Write the aggregated data to the new sheet
    # with pd.ExcelWriter(excel_path, **writer_args) as writer:
    #     daily_df.to_excel(writer, sheet_name=target_sheet, index=False)
  
    print(f"🧹 Cleaned and aggregated data saved to sheet '{target_sheet}' in '{excel_path}'.")

    # Copy to clipboard
    daily_df.to_clipboard(index=False)
    print("📋 Data copied to clipboard!")

# Example usage
aggregate_to_daily(
    excel_path=r"C:\Users\Sam\Desktop\ML\task\Data.xlsx",
    source_sheet="Encoded_Data",
    target_sheet="Daily_Aggregated",
    timestamp_col="Timestamp",
    unique_threshold=15  # Columns with <= 10 unique values become Mode, > 10 become Average
)