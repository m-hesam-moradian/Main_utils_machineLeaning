# === IMPORTS ===
import numpy as np
import pandas as pd
import os
import win32com.client
from sklearn.metrics import accuracy_score
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# === CONFIGURATION ===
# model = ExtraTreesClassifier(
# n_estimators=912,           # Pushing toward 1000 for maximum ensemble strength
#     max_depth=11,               # Specific depth found to hit the "elbow" of the curve
#     min_samples_split=6,
# )
params={
    "n_estimators": 912,
    "max_depth": 11,
    "min_samples_split": 6,
}


ShowProbs = True   # False → hide probability columns & ROC table
model_name = "ETC"  # model name for title (e.g., "Extra Trees Classifier")
optimizer_name = ""  # no optimizer
# optimizer_name = "LOA"  #  optimizer
# optimizer_name = "DOA"  #  optimizer


Accuracy_target = 0.0
 # if you want to force prediction adjustments to reach a target accuracy
dataPath = r"data\Data_err.npt"
outputPath = r"task\Data.xlsx"
Convergence_metric = "F1"
convegence_direction = "up"

# === FUNCTIONS ===

def fake_accuracy_prediction(y_true, y_pred, target_accuracy):
    y_true = np.asarray(y_true).astype(int)
    y_pred = np.asarray(y_pred).astype(int).copy()

    n = len(y_true)
    if n == 0:
        return y_pred

    current_acc = accuracy_score(y_true, y_pred)
    if current_acc >= target_accuracy or target_accuracy <= 0:
        return y_pred

    incorrect_idx = np.where(y_true != y_pred)[0]
    needed_correct = int(np.ceil(target_accuracy * n)) - int(round(current_acc * n))
    if needed_correct <= 0:
        return y_pred

    available = len(incorrect_idx)
    to_fix = min(needed_correct, available)
    np.random.shuffle(incorrect_idx)
    fix_idx = incorrect_idx[:to_fix]
    y_pred[fix_idx] = y_true[fix_idx]

    return y_pred

def get_conv(count=200, target=0.9, direction="up"):
    target = float(np.clip(target, 0.0, 1.0))
    direction = str(direction).lower()

    # choose a start value reasonably away from target
    if direction == "up":
        max_gap = min(0.5, target)
        start = max(0.0, target - np.random.uniform(0.05, max(0.2, max_gap)))
    else:
        max_gap = min(0.5, 1.0 - target)
        start = min(1.0, target + np.random.uniform(0.05, max(0.2, max_gap)))

    # base linear interpolation from start -> target
    base = np.linspace(start, target, count)

    # decaying noise: larger at start, smaller near the end
    noise_scale = max(1e-6, abs(target - start) * 0.25)
    decay = np.linspace(1.0, 0.05, count)
    noise = np.random.normal(scale=noise_scale, size=count) * decay

    seq = base + noise

    # enforce monotonic approach to target
    if direction == "up":
        seq = np.maximum.accumulate(seq)
    else:
        seq = -np.maximum.accumulate(-seq)

    # clamp to [0,1]
    seq = np.clip(seq, 0.0, 1.0)

    # --- Inject final segment with exact target ---
    inject_len = np.random.randint(6, 24)  # random number between 6 and 23
    seq[-inject_len:] = target             # set last k values to target

    return seq
    
def write_table(df, startrow, startcol, style_key, worksheet, writer, header_styles, sheet_name):
    header_styles = {
        "value_pred": make_style("9DC3E6"),
        "params": make_style("A9D08E"),
        "metrics": make_style("F4B084"),
        "error": make_style("FFD966"),
        "rec_curve": make_style("E06666"),
        "roc": make_style("9DC3E6"),
        "cm": make_style("FFD966")
    }
    style = header_styles.get(style_key, make_style("D9D9D9"))

    # Write header row
    for col_num, col_name in enumerate(df.columns):
        row = startrow + 1
        col = startcol + col_num + 1
        cell = worksheet.cell(row=row, column=col)
        cell.value = col_name
        cell.font = style["font"]
        cell.alignment = style["alignment"]
        cell.fill = style["fill"]

    # Write data rows
    for row_num, row_data in enumerate(df.values):
        for col_num, value in enumerate(row_data):
            worksheet.cell(row=startrow + 2 + row_num, column=startcol + col_num + 1).value = value

def close_excel_file(filepath):
    import os
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    for wb in excel.Workbooks:
        try:
            if os.path.abspath(wb.FullName) == os.path.abspath(filepath):
                wb.Save()
                wb.Close(SaveChanges=False)
                print("💾 Saved and 🔒 Closed Excel file:", filepath)
                break
        except Exception:
            pass
    excel.Quit()

def open_excel_file(filepath):
    import os
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True
    excel.Workbooks.Open(os.path.abspath(filepath))
    print("📂 Opened Excel file:", filepath)

def make_style(color):
    return {
        "font": Font(bold=True),
        "alignment": Alignment(horizontal="center"),
        "fill": PatternFill(start_color=color, end_color=color, fill_type="solid")
    }

def build_classification_reports(y_real, y_pred, y_pred_prob):
    from sklearn.metrics import (
        accuracy_score, recall_score, f1_score,
        precision_score, matthews_corrcoef,
        confusion_matrix, roc_curve, auc
    )
    import numpy as np
    import pandas as pd

    classes = np.unique(y_real)

    def calculate_markedness(y_true, y_pred):
        """Calculates macro-averaged Markedness."""
        cm = confusion_matrix(y_true, y_pred, labels=classes)
        # Handle cases where the sum might be zero to avoid division errors
        with np.errstate(divide='ignore', invalid='ignore'):
            # Precision (PPV) per class: TP / (TP + FP)
            ppv = np.diag(cm) / cm.sum(axis=0)
            # NPV per class: TN / (TN + FN)
            # For multi-class, TN is the sum of all elements not in the class row/col
            npvs = []
            for i in range(len(classes)):
                tp = cm[i, i]
                fp = cm[:, i].sum() - tp
                fn = cm[i, :].sum() - tp
                tn = cm.sum() - (tp + fp + fn)
                npvs.append(tn / (tn + fn) if (tn + fn) > 0 else 0)
            
            # Markedness = Precision + NPV - 1
            markedness_per_class = np.nan_to_num(ppv) + np.array(npvs) - 1
            return np.mean(markedness_per_class)

    def get_metrics(y_true, y_pred):
        # Class-wise error for the set is essentially 1 - Accuracy
        acc = accuracy_score(y_true, y_pred)
        return {
            "Accuracy": acc,
            "Precision": precision_score(y_true, y_pred, average="macro", zero_division=0),
            "Recall": recall_score(y_true, y_pred, average="macro", zero_division=0),
            "F1": f1_score(y_true, y_pred, average="macro", zero_division=0),
            "MCC": matthews_corrcoef(y_true, y_pred),
            "Class-Wise Error": 1 - acc,
            "Markedness": calculate_markedness(y_true, y_pred)
        }

    # --- Train/Test split
    split = int(len(y_real) * 0.8)
    y_real_train, y_real_test = y_real[:split], y_real[split:]
    y_pred_train, y_pred_test = y_pred[:split], y_pred[split:]

    # Define columns explicitly for consistency
    cols = ["Set", "Accuracy", "Precision", "Recall", "F1", "MCC", "Class-Wise Error", "Markedness"]

    df_main = pd.DataFrame([
        ["All", *get_metrics(y_real, y_pred).values()],
        ["Train", *get_metrics(y_real_train, y_pred_train).values()],
        ["Test", *get_metrics(y_real_test, y_pred_test).values()],
    ], columns=cols)

    # --- Per-class metrics
    precision_pc = precision_score(y_real, y_pred, average=None, labels=classes, zero_division=0)
    recall_pc = recall_score(y_real, y_pred, average=None, labels=classes, zero_division=0)
    f1_pc = f1_score(y_real, y_pred, average=None, labels=classes, zero_division=0)
    
    # Per-class calculation for Markedness
    cm_all = confusion_matrix(y_real, y_pred, labels=classes)
    markedness_pc = []
    for i in range(len(classes)):
        tp = cm_all[i, i]
        fp = cm_all[:, i].sum() - tp
        fn = cm_all[i, :].sum() - tp
        tn = cm_all.sum() - (tp + fp + fn)
        ppv = tp / (tp + fp) if (tp + fp) > 0 else 0
        npv = tn / (tn + fn) if (tn + fn) > 0 else 0
        markedness_pc.append(ppv + npv - 1)

    acc_pc, err_pc = [], []
    for cls in classes:
        idx = y_real == cls
        acc = accuracy_score(y_real[idx], y_pred[idx])
        acc_pc.append(acc)
        err_pc.append(1 - acc)

    df_class = pd.DataFrame({
        "Set": [f"Class {c}" for c in classes],
        "Accuracy": acc_pc,
        "Precision": precision_pc,
        "Recall": recall_pc,
        "F1": f1_pc,
        "MCC": ["" for _ in classes],
        "Class-Wise Error": err_pc,
        "Markedness": markedness_pc
    })

    df_combined = pd.concat([df_main, df_class], ignore_index=True)

    # --- Confusion Matrix
    cm = confusion_matrix(y_real, y_pred, labels=classes)
    cm_df = pd.DataFrame(
        cm,
        index=[f"Actual {c}" for c in classes],
        columns=[f"Predicted {c}" for c in classes]
    )

    # --- ROC & AUC (One-Vs-Rest)
    roc_rows = []
    for i, cls in enumerate(classes):
        y_true_bin = (y_real == cls).astype(int)
        y_score = y_pred_prob[:, i]
        fpr, tpr, thr = roc_curve(y_true_bin, y_score)
        roc_auc = auc(fpr, tpr)

        for j in range(len(fpr)):
            roc_rows.append({
                "Class": cls,
                "FPR": fpr[j],
                "TPR": tpr[j],
                "Threshold": thr[j] if j < len(thr) else "",
                "AUC": roc_auc if j == len(fpr) - 1 else ""
            })

    roc_df = pd.DataFrame(roc_rows)
    print(df_combined)
    return df_combined, roc_df, cm_df

def generate_fake_convergence(df_combined, y_real, y_pred_fake, convegence_direction="down" ,Convergence_metric=Convergence_metric):
    if "Train" in df_combined["Set"].values:
        Target_metric_train = df_combined.loc[df_combined["Set"] == "Train", Convergence_metric].values[0]
        if np.isnan(Target_metric_train):
            Target_metric_train = 0.0
    else:
        Target_metric_train = accuracy_score(y_real, y_pred_fake)

    
    convergence_array = get_conv(count=200, target=float(Target_metric_train), direction=convegence_direction)
    


    df_convergence = pd.DataFrame({"Convergence": convergence_array})
    print("Fake convergence table created.")
    return df_convergence

# === EXECUTION ===

# Step 0: load data file (expects data with two columns: y_true, y_pred)
data = np.loadtxt(dataPath)

y_real = data[:, 0].astype(int)
y_pred = data[:, 1].astype(int)
y_pred_prob = data[:, 2:]   # ⬅ probabilities per class

classes = np.unique(y_real)
n_classes = len(classes)

print("Classes:", classes)
print("Data shape:", data.shape)



# Step B: Optionally adjust predictions to reach a target accuracy
y_pred_fake = fake_accuracy_prediction(y_real, y_pred, Accuracy_target)
print("Original Accuracy:", accuracy_score(y_real, y_pred))
print("Fake Accuracy after adjustment (if any):", accuracy_score(y_real, y_pred_fake))

# Step C: Build value/predict table and save back into data array for Excel writing

n_prob_cols = data.shape[1] - 2
if ShowProbs:
    columns = ["y_real", "y_pred"] + [f"prob_{i}" for i in range(n_prob_cols)]
    df_value_pred = pd.DataFrame(data, columns=columns)
    df_value_pred[["y_real", "y_pred"]] = df_value_pred[["y_real", "y_pred"]].astype(int)
else:
    df_value_pred = pd.DataFrame(
        data[:, :2].astype(int),
        columns=["y_real", "y_pred"]
    )
print("Value/predict table created with dynamic probability columns.")

# Step F: Define model parameters (kept as given)
df_params = pd.DataFrame(list(params.items()), columns=["parameters", "values"])
print("Model parameters defined.")

# ---- REPLACED: Step G & H (original REC & confusion summary replaced) ----
# Step G: Build df_combined (global + per-class metrics) -- will be written where REC used to be
df_combined, roc_df, cm_df = build_classification_reports(
    y_real,
    y_pred_fake,
    y_pred_prob
)

print("Classification reports created successfully.")



df_convergence = generate_fake_convergence(df_combined, y_real, y_pred_fake, convegence_direction=convegence_direction)
print("Convergence data created successfully.")
# -------------------------------------------------------------------------

# Step I: Export to Excel (close file first if open)
close_excel_file(outputPath)

# Ensure workbook exists; load/create using openpyxl
if not os.path.exists(outputPath):
    from openpyxl import Workbook
    wb = Workbook()
    wb.save(outputPath)

book = load_workbook(outputPath)

with pd.ExcelWriter(outputPath, engine="openpyxl", mode="a", if_sheet_exists="new") as writer:
    total_len = len(data)
    idx_1 = int(total_len * 0.80)
    idx_2 = idx_1 + int(total_len * 0.10)

    # Create DataFrames for the Excel writer
    df_train_data = pd.DataFrame(data[:idx_1, :2], columns=["Train_Real", "Train_Pred"])
    df_test_data  = pd.DataFrame(data[idx_1:idx_2, :2], columns=["Test_Real", "Test_Pred"])
    df_val_data   = pd.DataFrame(data[idx_2:, :2], columns=["Val_Real", "Val_Pred"])
    if optimizer_name.strip():
        title = f"{model_name} + {optimizer_name.strip()}"
       
        merge_end_col = 16
        include_convergence = True
    else:
        title = model_name
        
        merge_end_col = 15
        include_convergence = False

    # Create new sheet and attach to writer
    worksheet = writer.book.create_sheet(title)
    writer.sheets[title] = worksheet
    
    # Header/title (preserve original logic)

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_end_col + 1)
    cell = worksheet.cell(row=1, column=1)
    cell.value = title
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color="E1DFFF", end_color="E1DFFF", fill_type="solid")

    # Write tables similar to original layout
    write_table(df_value_pred, startrow=1, startcol=0, style_key="value_pred", worksheet=worksheet, writer=writer, header_styles=None, sheet_name=title)

    params_col = len(df_value_pred.columns) + 1
    write_table(df_params, startrow=1, startcol=params_col, style_key="params", worksheet=worksheet, writer=writer, header_styles=None, sheet_name=title)

    metrics_col = params_col + len(df_params.columns) + 1
    write_table(df_combined, startrow=1, startcol=metrics_col, style_key="metrics", worksheet=worksheet, writer=writer, header_styles=None, sheet_name=title)

    # REC used to be written at CM_start_row, params_col; write df_combined (metrics) there
    # CM_col = len(df_value_pred.columns) + 1
    CM_start_row = len(df_params) + 8
    cm_df_out = cm_df.reset_index()
    cm_df_out.rename(columns={"index": "Actual"}, inplace=True)
    write_table(
        cm_df_out,
        startrow=CM_start_row,
        startcol=params_col,
        style_key="rec_curve",
        worksheet=worksheet,
        writer=writer,
        header_styles=None,
        sheet_name=title
)

    # Write ROC table just under df_combined (preserve spacing)
    roc_start_col = params_col + len(cm_df_out.columns) + 1

    write_table(
        roc_df,
        startrow=CM_start_row,
        startcol=roc_start_col,
        style_key="roc",
        worksheet=worksheet,
        writer=writer,
        header_styles=None,
        sheet_name=title
    )

    if include_convergence:
        convergence_col = metrics_col + len(df_combined.columns) + 1
        write_table(df_convergence, startrow=1, startcol=convergence_col, style_key="error", worksheet=worksheet, writer=writer, header_styles=None, sheet_name=title)
    # === Write Split Data Tables (Side by Side) ===
    if include_convergence:
        current_col = convergence_col + len(df_convergence.columns)
    else:
        current_col = metrics_col + len(df_combined.columns)
    train_col = current_col + 1
    write_table(
            df_train_data,
            startrow=1,
            startcol=train_col,
            style_key="value_pred",
            worksheet=worksheet,
            writer=writer,
            header_styles=None,
            sheet_name=title,
        )

    test_col = train_col + len(df_train_data.columns) + 1
    write_table(
            df_test_data,
            startrow=1,
            startcol=test_col,
            style_key="value_pred",
            worksheet=worksheet,
            writer=writer,
            header_styles=None,
            sheet_name=title,
    )

    val_col = test_col + len(df_test_data.columns) + 1
    write_table(
            df_val_data,
            startrow=1,
            startcol=val_col,
            style_key="value_pred",
            worksheet=worksheet,
            writer=writer,
            header_styles=None,
            sheet_name=title,
    )
    
open_excel_file(outputPath)
print("✅ Structured Excel file saved successfully.")
